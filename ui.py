"""
ui.py - Flask web interface for the amortization estimation program.

Run with:
    python ui.py
Then open: http://localhost:5000
"""
# NOTE: index.html uses SSE streaming (/api/stream/<job_id>).
# dashboard.html and economy.html use the legacy /api/status polling endpoint.
# Both coexist in this file.

import os
import sys
import uuid
import queue
import threading
from datetime import datetime
from flask import Flask, render_template, jsonify, request, send_file, Response, stream_with_context, session, redirect, url_for

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

import yaml
from generator import generate
from db import get_connection

app = Flask(__name__, template_folder=os.path.join(SCRIPT_DIR, "templates"))
app.secret_key = "hasnaoui-amortissement-secret-key-123456"

# ── Legacy global job state (used by dashboard / /api/status polling) ────────
_lock = threading.Lock()
_job = {
    "status": "idle",   # idle | running | done | error
    "logs": [],
    "output_file": None,
    "row_counts": {},
    "started_at": None,
    "finished_at": None,
}

# ── Per-job SSE store (used by the generator index page) ─────────────────────
# job_id → {"status": str, "log_q": Queue, "file": str|None}
_sse_jobs: dict[str, dict] = {}
_sse_lock = threading.Lock()


def _log(msg: str):
    ts = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    with _lock:
        _job["logs"].append(line)
    print(line)


def _load_config():
    path = os.path.join(SCRIPT_DIR, "config.yaml")
    with open(path, encoding="utf-8") as f:
        cfg = yaml.safe_load(f) or {}
    
    cats = cfg.get("categories", [])
    if not isinstance(cats, list):
        cats = []
    
    has_virtual = any(c.get("is_virtual") for c in cats)
    if not has_virtual:
        cats.extend([
            {"ids": [-1], "label": "Bâtiments d'habitation", "years": 50, "is_virtual": True, "asset_ids": [245131, 245132, 245133], "default_years": 50.0},
            {"ids": [-2], "label": "PRESSES ET COMPRESSEUR", "years": 10, "is_virtual": True, "asset_ids": [246689, 247684, 247904], "default_years": 10.0}
        ])
        cfg["categories"] = cats
        
    return cfg



def _parse_year_debut(val_s):
    """Parse year_debut which can be a 4-digit year or a date (DD/MM/YYYY or YYYY-MM-DD)."""
    if not val_s:
        return None, None
    val_s = str(val_s).strip()
    if not val_s:
        return None, None

    from datetime import datetime as _dt
    # Check DD/MM/YYYY
    if "/" in val_s:
        parts = val_s.split("/")
        if len(parts) == 3:
            try:
                dt = _dt.strptime(val_s, "%d/%m/%Y").date()
                return dt, "date"
            except ValueError:
                pass

    # Check YYYY-MM-DD
    if "-" in val_s and len(val_s) == 10:
        try:
            dt = _dt.strptime(val_s, "%Y-%m-%d").date()
            return dt, "date"
        except ValueError:
            pass

    # Check YYYY (4 digits)
    if val_s.isdigit() and len(val_s) == 4:
        try:
            return int(val_s), "year"
        except ValueError:
            pass

    return None, None


def _parse_flexible_date(d_str: str):
    """Parse flexible date input (YYYY, MM/YYYY, DD/MM/YYYY, YYYY-MM-DD) into (str_formatted, date_obj)."""
    if not d_str or not str(d_str).strip():
        return None, None
    d_str = str(d_str).strip()
    import calendar
    from datetime import datetime as _dt, date as _date

    # Check YYYY-MM-DD
    if "-" in d_str and len(d_str) == 10:
        try:
            dt = _dt.strptime(d_str, "%Y-%m-%d").date()
            return d_str, dt
        except ValueError:
            pass

    # Check DD/MM/YYYY
    parts = d_str.split("/")
    if len(parts) == 3:
        try:
            dt = _dt.strptime(d_str, "%d/%m/%Y").date()
            return dt.strftime("%Y-%m-%d"), dt
        except ValueError:
            pass

    # Check MM/YYYY
    if len(parts) == 2:
        try:
            m, y = int(parts[0]), int(parts[1])
            last_d = calendar.monthrange(y, m)[1]
            dt = _date(y, m, last_d)
            return dt.strftime("%Y-%m-%d"), dt
        except Exception:
            pass

    # Check YYYY
    if d_str.isdigit() and len(d_str) == 4:
        try:
            y = int(d_str)
            dt = _date(y, 12, 31)
            return dt.strftime("%Y-%m-%d"), dt
        except Exception:
            pass

    return None, None


from data_cache import global_cache



def _run_generation(categories, output_format, asset_id=None, societe=None, dep_date=None,
                    granularity="monthly", from_date=None, as_of_date=None, sse_job_id=None, year_debut=None,
                    asset_search=None, non_amortie=False, year_debut_type="start_date", is_medina=False, ignore_overrides=False,
                    vnc_filter="all"):
    """
    Background thread: uses generator.py for batched processing.
    If sse_job_id is given, logs are also pushed to the SSE queue for that job.
    """
    with _lock:
        _job["status"] = "running"
        _job["logs"] = []
        _job["output_file"] = None
        _job["row_counts"] = {}
        _job["started_at"] = datetime.now().isoformat()
        _job["finished_at"] = None
    def combined_log(msg):
        _log(msg)
        if sse_job_id:
            with _sse_lock:
                j = _sse_jobs.get(sse_job_id)
            if j:
                j["log_q"].put(msg)
 
    try:
        combined_log("Initialisation de la generation...")
        cfg_path = os.path.join(SCRIPT_DIR, "config.yaml")
        with open(cfg_path, encoding="utf-8") as f:
            cfg = yaml.safe_load(f)
 
        raw_out = cfg.get("output", {}).get("directory", "output")
        out_dir = raw_out if os.path.isabs(raw_out) else os.path.join(SCRIPT_DIR, raw_out)
        fp = generate(
            categories=categories,
            output_format=output_format,
            output_dir=out_dir,
            asset_id=asset_id,
            societe=societe,
            dep_date=dep_date,
            granularity=granularity,
            from_date=from_date,
            as_of_date=as_of_date,
            log_fn=combined_log,
            year_debut=year_debut,
            year_debut_type=year_debut_type,
            asset_search=asset_search,
            non_amortie=non_amortie,
            is_medina=is_medina,
            ignore_overrides=ignore_overrides,
            vnc_filter=vnc_filter,
        )
    except Exception as e:
        combined_log(f"[ERREUR FATALE] {e}")
        import traceback
        print(traceback.format_exc())
        fp = None

    with _lock:
        _job["output_file"] = fp
        _job["status"] = "done" if fp else "error"
        _job["finished_at"] = datetime.now().isoformat()

    if sse_job_id:
        with _sse_lock:
            j = _sse_jobs.get(sse_job_id)
        if j:
            j["file"]   = fp
            j["status"] = "done" if fp else "error"
            j["log_q"].put(None)   # sentinel — stream ends


def _get_virtual_category_mapping(all_cats):
    """
    Returns (v_map, v_aids_map)
    v_map: asset_id -> virtual_cat_id
    v_aids_map: virtual_cat_id -> list of asset_ids
    """
    v_map = {}
    v_aids_map = {}
    
    # Add defaults if no virtual categories are present
    has_virtual = any(c.get("is_virtual") for c in all_cats)
    cats_to_parse = list(all_cats)
    if not has_virtual:
        cats_to_parse.extend([
            {"ids": [-1], "label": "Bâtiments d'habitation", "years": 50, "is_virtual": True, "asset_ids": [245131, 245132, 245133]},
            {"ids": [-2], "label": "PRESSES ET COMPRESSEUR", "years": 10, "is_virtual": True, "asset_ids": [246689, 247684, 247904]}
        ])
        
    for cat in cats_to_parse:
        if cat.get("is_virtual"):
            v_id = cat["ids"][0]
            aids = [int(aid) for aid in cat.get("asset_ids", [])]
            v_aids_map[v_id] = aids
            for aid in aids:
                v_map[aid] = v_id
    return v_map, v_aids_map


_excel_override_cache = None

def _get_excel_category_overrides():
    """
    Load Classeur1.xlsx and return a dict {id_immo (int) -> new_cat_id (int)}.
    Results are cached in-process.
    """
    global _excel_override_cache
    if _excel_override_cache is not None:
        return _excel_override_cache
    try:
        import pandas as pd
        xl_path = os.path.join(os.path.dirname(__file__), "Classeur1.xlsx")
        df = pd.read_excel(xl_path)
        # Normalise column names: tolower + strip
        df.columns = [c.lower().strip() for c in df.columns]
        id_col  = next((c for c in df.columns if "immo" in c), None)
        cat_col = next((c for c in df.columns if "cat" in c), None)
        if id_col and cat_col:
            mapping = {
                int(row[id_col]): int(row[cat_col])
                for _, row in df.iterrows()
                if pd.notna(row[id_col]) and pd.notna(row[cat_col])
            }
            _excel_override_cache = mapping
            return mapping
    except Exception as e:
        print(f"[WARN] Could not load Classeur1.xlsx: {e}")
    _excel_override_cache = {}
    return {}


# ── Authentication Middleware & Routes ────────────────────────────────────────

# ── Authentication Middleware & Routes ────────────────────────────────────────

@app.before_request
def check_auth():
    if request.endpoint in ('login', 'static'):
        return
    if 'username' not in session:
        return redirect(url_for('login'))
    if session.get('role') == 'user':
        if request.endpoint in ('config_page', 'api_update_config', 'economy_page'):
            return redirect(url_for('index'))


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        if username == "admin" and password == "admin":
            session["username"] = "admin"
            session["role"] = "admin"
            return redirect(url_for("index"))
        elif username == "user" and password == "user":
            session["username"] = "user"
            session["role"] = "user"
            return redirect(url_for("index"))
        else:
            error = "Identifiants incorrects (admin/admin ou user/user)"
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/config")
def config_page():
    return render_template("config.html")


@app.route("/api/config")
def api_config():
    """Return config info (DB host/dbname + categories). No password."""
    try:
        cfg = _load_config()
        db = cfg["database"]
        all_cats = cfg.get("categories", [])

        # Add defaults if no virtual categories are present
        has_virtual = any(c.get("is_virtual") for c in all_cats)
        if not has_virtual:
            all_cats.extend([
                {"ids": [-1], "label": "Bâtiments d'habitation", "years": 50, "is_virtual": True, "asset_ids": [245131, 245132, 245133]},
                {"ids": [-2], "label": "PRESSES ET COMPRESSEUR", "years": 10, "is_virtual": True, "asset_ids": [246689, 247684, 247904]}
            ])
            
        real_cats = [c for c in all_cats if not c.get("is_virtual")]
        virtual_cats = [c for c in all_cats if c.get("is_virtual")]

        # Fetch old_years from DB
        try:
            conn = get_connection(db)
            all_cids = []
            for c in real_cats:
                all_cids.extend(c.get("ids", []))
            if all_cids:
                with conn.cursor() as cur:
                    cur.execute("""
                        SELECT id, method_number
                        FROM account_asset_category
                        WHERE id = ANY(%s)
                    """, [all_cids])
                    m_map = {r[0]: r[1] for r in cur.fetchall()}
                for c in real_cats:
                    cids = c.get("ids", [])
                    if cids and cids[0] in m_map and m_map[cids[0]] is not None:
                        c["old_years"] = round(float(m_map[cids[0]]) / 12.0, 2)
                    else:
                        # null method_number = non-amortizable in Odoo → show None (displayed as '—')
                        c["old_years"] = None
            conn.close()
        except Exception as ex:
            # Fallback if DB is not reachable
            for c in real_cats:
                c["old_years"] = c.get("years") or None

        for c in real_cats:
            c["is_virtual"] = False
            c["asset_ids"] = []

        for c in virtual_cats:
            c["is_virtual"] = True
            # old_years for virtual = their configured years (they have no DB method_number)
            c["old_years"] = c.get("years", 10.0)
            if "asset_ids" not in c:
                c["asset_ids"] = []

        combined_cats = real_cats + virtual_cats

        return jsonify({
            "db": {
                "host": db.get("host"),
                "port": db.get("port", 5432),
                "dbname": db.get("dbname"),
                "user": db.get("user"),
            },
            "categories": combined_cats,
            "output_format": cfg.get("output", {}).get("format", "excel"),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/config", methods=["POST"])
def api_update_config():
    """Update config from web."""
    try:
        payload = request.json or {}
        categories_data = payload.get("categories", [])
        
        # Clean up temporary fields before saving
        for c in categories_data:
            c.pop("old_years", None)
            if "asset_ids" in c:
                c["asset_ids"] = [int(x) for x in c["asset_ids"] if str(x).strip().isdigit()]
        
        cfg = _load_config()
        cfg["categories"] = categories_data
        
        cfg_path = os.path.join(SCRIPT_DIR, "config.yaml")
        with open(cfg_path, "w", encoding="utf-8") as f:
            yaml.dump(cfg, f, default_flow_style=False, allow_unicode=True)
            
        # Re-trigger data cache refresh in the background
        from data_cache import global_cache
        threading.Thread(target=global_cache.refresh, args=(cfg,), daemon=True).start()
        
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


@app.route("/economy")
def economy_page():
    return render_template("economy.html")


@app.route("/api/test-connection")
def api_test_connection():
    """Quick DB ping."""
    try:
        cfg = _load_config()
        conn = get_connection(cfg["database"])
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 200


@app.route("/api/filters")
def api_filters():
    """Return available dates (from depreciation table) and companies."""
    try:
        with global_cache.lock:
            cached_dates = global_cache.dates
            cached_socs = global_cache.societes
            
        if cached_dates:
            return jsonify({
                "dates": cached_dates,
                "societes": cached_socs
            })

        cfg = _load_config()
        conn = get_connection(cfg["database"])
        result = {}
        with conn.cursor() as cur:
            cur.execute("""
                SELECT DISTINCT depreciation_date
                FROM account_asset_depreciation_line
                WHERE depreciation_date >= '2022-01-01'
                ORDER BY depreciation_date DESC
                LIMIT 50
            """)
            result["dates"] = [r[0].isoformat() if r[0] else None for r in cur.fetchall()]

            cur.execute("""
                SELECT DISTINCT rc.name
                FROM res_company rc
                JOIN account_asset_asset aaa ON aaa.company_id = rc.id
                WHERE rc.name IS NOT NULL
                ORDER BY rc.name
            """)
            result["societes"] = [r[0] for r in cur.fetchall()]

        conn.close()
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/dashboard")
def dashboard():
    return render_template("dashboard.html")


@app.route("/company_details")
def company_details():
    return render_template("company_details.html")


@app.route("/api/dashboard")
def api_dashboard():
    """Dashboard: gap réel vs estimé, assets by acq year, monthly evolution."""
    try:
        import psycopg2.extras
        from generator import _compute_schedule
        from datetime import datetime as dt_cls, date as date_cls

        date_str_raw = request.args.get("date") or None
        date_str, target_date = _parse_flexible_date(date_str_raw)
        societe      = request.args.get("societe") or None
        if societe in ["Toutes", "undefined", "null", "all"]:
            societe = None
        gran         = request.args.get("gran") or "monthly"
        cat_ids_s    = request.args.get("cat_ids") or ""
        asset_id_f   = request.args.get("asset_id") or None   # filter by ID
        asset_search = (request.args.get("asset_search") or "").strip()  # name/code
        year_debut_s = request.args.get("year_debut") or None
        yd_val, yd_type = _parse_year_debut(year_debut_s)
        year_debut_type = request.args.get("year_debut_type") or "start_date"
        non_amortie = request.args.get("non_amortie") == "true"
        vnc_filter   = request.args.get("vnc_filter") or "all"
        if vnc_filter == "positive":
            non_amortie = True

        if target_date:
            if date_str_raw and len(date_str_raw.strip()) == 4 and date_str_raw.strip().isdigit():
                eff_gran = "yearly"
            else:
                eff_gran = gran

            if eff_gran == "yearly":
                first_day_str = f"{target_date.year}-01-01"
                last_day_str  = f"{target_date.year}-12-31"
                first_day = date_cls(target_date.year, 1, 1)
                last_day  = date_cls(target_date.year, 12, 31)
            elif eff_gran == "monthly":
                import calendar
                first_day = target_date.replace(day=1)
                last_day = target_date.replace(day=calendar.monthrange(target_date.year, target_date.month)[1])
                first_day_str = first_day.strftime("%Y-%m-%d")
                last_day_str  = last_day.strftime("%Y-%m-%d")
            else: # daily
                first_day = target_date
                last_day = target_date
                first_day_str = date_str
                last_day_str  = date_str

        cfg      = _load_config()
        all_cats = cfg.get("categories", [])

        if cat_ids_s:
            sel_ids = set(int(x) for x in cat_ids_s.split(",") if x.strip())
            cats = [c for c in all_cats if any(cid in sel_ids for cid in c["ids"])]
        else:
            cats = all_cats

        cat_lookup = {}
        for cat in cats:
            for cid in cat["ids"]:
                _yrs = int(cat["years"]) if "years" in cat else (int(cat["months"]) / 12 if "months" in cat else 10)
                cat_lookup[cid] = {"label": cat.get("label", "?"),
                                   "years": _yrs,
                                   "old_years": cat.get("old_years"),
                                   "is_virtual": cat.get("is_virtual", False)}
        all_cat_ids = list(cat_lookup.keys())
        if not all_cat_ids:
            return jsonify({"error": "Aucune categorie."}), 400

        # Check cache
        with global_cache.lock:
            cached_assets = global_cache.assets
            cached_lines = global_cache.lines

        use_cache = bool(cached_assets)

        sc = "AND societe_comptable.name = %s" if societe else ""
        af = "AND aaat.id = %s" if asset_id_f else ""
        sf = "AND (aaat.name ILIKE %s OR aaat.code ILIKE %s)" if asset_search else ""
        date_col = "aaat.purchase_date" if year_debut_type == "purchase_date" else "aaat.date_comptabilisation"
        if yd_type == "date":
            yf = f"AND {date_col} = %s"
        elif yd_type == "year":
            yf = f"AND EXTRACT(YEAR FROM {date_col}) = %s"
        else:
            yf = ""
        conn = get_connection(cfg["database"])

        # ── 0. Fetch old years (method_number) per category from DB ───────────
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, method_number
                FROM account_asset_category
                WHERE id = ANY(%s)
            """, [all_cat_ids])
            method_map = {r[0]: float(r[1]) if r[1] is not None else 0.0 for r in cur.fetchall()}  # cat_id -> method_number
        # old_years = method_number / 12
        for cid in cat_lookup:
            if cat_lookup[cid].get("is_virtual"):
                cat_lookup[cid]["old_years"] = cat_lookup[cid]["years"]
            else:
                mn = method_map.get(cid, 0)
                cat_lookup[cid]["old_years"] = round(mn / 12, 2) if mn else None

        if use_cache:
            # Filter cached_assets in Python
            asset_rows = []
            sel_cat_set = set(all_cat_ids)
            for a in cached_assets:
                if a["cat_id"] not in sel_cat_set: continue
                if societe and a["company_name"] != societe: continue
                if asset_id_f and str(a["id_immo"]) != str(asset_id_f): continue
                if yd_val is not None:
                    try:
                        sd = a.get(year_debut_type)
                        if isinstance(sd, str):
                            sd_parsed = dt_cls.strptime(sd[:10], "%Y-%m-%d").date()
                        elif isinstance(sd, (date_cls, dt_cls)):
                            sd_parsed = sd if isinstance(sd, date_cls) else sd.date()
                        else:
                            sd_parsed = None
                        
                        if yd_type == "date":
                            if sd_parsed != yd_val: continue
                        elif yd_type == "year":
                            if sd_parsed.year != yd_val: continue
                    except: continue
                if asset_search:
                    s = asset_search.lower()
                    if s not in (a["code"] or "").lower() and s not in (a["asset_name"] or "").lower(): continue
                
                # VNC filter (REAL-based)
                is_close = a.get("state") == "close"
                a_lines = cached_lines.get(a["id_immo"], [])
                real_cum = 0.0
                if a_lines:
                    relevant = [l for l in a_lines if (not date_str or l[0] <= date_str)]
                    if relevant:
                        last_line = sorted(relevant, key=lambda x: x[0])[-1]
                        real_cum = last_line[2] + last_line[1]
                is_fully_amortized = is_close or (real_cum >= a["purchase_value"] - 0.01)

                if vnc_filter == "positive" or non_amortie:
                    if is_fully_amortized:
                        continue
                elif vnc_filter == "zero":
                    if not is_fully_amortized:
                        continue
                
                asset_rows.append(a)
                
            asset_ids = [a["id_immo"] for a in asset_rows]
            
            if date_str:
                date_lines = {}
                for aid in asset_ids:
                    valid_lines = [item for item in cached_lines.get(aid, []) if item[0] <= date_str]
                    cum_real = 0.0
                    if valid_lines:
                        valid_lines.sort(key=lambda x: x[0])
                        cum_real = valid_lines[-1][2] + valid_lines[-1][1]
                    
                    period_real = sum(item[1] for item in cached_lines.get(aid, []) if first_day_str <= item[0] <= last_day_str)
                    date_lines[aid] = (float(cum_real), float(period_real))
            else:
                real_by_cat_id = {}
                real_by_company = {}
                real_by_comp_cat_id = {}
                for a in asset_rows:
                    aid = a["id_immo"]
                    cid = a["cat_id"]
                    comp = a["company_name"]
                    total_real = sum(float(amt) for _d, amt, _dep in cached_lines.get(aid, []))
                    real_by_cat_id[cid] = real_by_cat_id.get(cid, 0.0) + total_real
                    real_by_company[comp] = real_by_company.get(comp, 0.0) + total_real
                    real_by_comp_cat_id[(comp, cid)] = real_by_comp_cat_id.get((comp, cid), 0.0) + total_real
            lines = []
        else:
            # ── 1. Assets (minimal) ───────────────────────────────────────────────
            v_map, v_aids_map = _get_virtual_category_mapping(all_cats)
            excel_map = _get_excel_category_overrides()  # id_immo -> new real cat_id
            v_aids = []
            for cid in all_cat_ids:
                if cid in v_aids_map:
                    v_aids.extend(v_aids_map[cid])
            # Also include assets whose Excel override points to one of the selected categories
            excel_aids = [aid for aid, cid in excel_map.items() if cid in all_cat_ids]
            all_extra_aids = list(set(v_aids + excel_aids))
            v_clause = f"OR aaat.id = ANY(ARRAY{all_extra_aids})" if all_extra_aids else ""

            extra_params = [all_cat_ids]
            if societe:      extra_params.append(societe)
            if asset_id_f:   extra_params.append(int(asset_id_f))
            if asset_search: extra_params += [f"%{asset_search}%", f"%{asset_search}%"]
            if yd_val is not None: extra_params.append(yd_val)

            with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
                cur.execute(f"""
                    SELECT aaat.id AS id_immo,
                           aaat.code,
                           aaat.name AS asset_name,
                           aaat.purchase_date,
                           aaat.date_comptabilisation AS start_date,
                            ROUND(
                                (
                                    (COALESCE(aaat.unit_price, 0) * COALESCE(aaat.quantite, 0))
                                    + COALESCE(aaat.costs, 0)
                                )::numeric,
                                2
                            ) AS purchase_value,
                           aaat.method_number,
                           aac.id AS cat_id,
                           COALESCE(societe_comptable.name, 'Non défini') AS company_name,
                           aaat.state
                    FROM account_asset_asset aaat
                    JOIN account_asset_category aac ON aac.id = aaat.category_id
                    LEFT JOIN res_company societe_comptable ON societe_comptable.id = aaat.company_id
                    WHERE (aac.id = ANY(%s) {v_clause}) {sc} {af} {sf} {yf}
                """, extra_params)
                
                asset_rows = []
                for r in cur.fetchall():
                    d = dict(r)
                    aid = d["id_immo"]
                    # Excel override takes priority over virtual map
                    if aid in excel_map:
                        d["cat_id"] = excel_map[aid]
                    elif aid in v_map:
                        d["cat_id"] = v_map[aid]
                    
                    if d["cat_id"] not in all_cat_ids:
                        continue
                    
                    # Apply VNC filter (pass 1: check 'close' state)
                    is_close = d.get("state") == "close"
                    if vnc_filter == "positive" or non_amortie:
                        if is_close:
                            continue
                    elif vnc_filter == "zero":
                        # If date_str is present, we will do a second pass check with exact cumulative real.
                        # If date_str is not present, we rely on 'close' state as a proxy.
                        if not date_str and not is_close:
                            continue
                        
                    asset_rows.append(d)

            asset_ids = [r["id_immo"] for r in asset_rows]

            # ── 2. Gap data: different strategy depending on whether a date is selected ──
            if date_str:
                if asset_ids:
                    with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
                        cur.execute(f"""
                            SELECT DISTINCT ON (asset_id) asset_id, depreciated_value, amount
                            FROM account_asset_depreciation_line
                            WHERE asset_id = ANY(%s) AND depreciation_date <= %s
                            ORDER BY asset_id, depreciation_date DESC
                        """, [asset_ids, date_str])
                        # Cast to float to avoid TypeError with Decimal
                        cum_map = {r["asset_id"]: float(r["depreciated_value"] or 0) + float(r["amount"] or 0) for r in cur.fetchall()}
                        
                        cur.execute(f"""
                            SELECT asset_id, SUM(amount) AS period_amt
                            FROM account_asset_depreciation_line
                            WHERE asset_id = ANY(%s) AND depreciation_date BETWEEN %s AND %s
                            GROUP BY asset_id
                        """, [asset_ids, first_day_str, last_day_str])
                        period_map = {r["asset_id"]: float(r["period_amt"] or 0) for r in cur.fetchall()}
                        
                        date_lines = {aid: (cum_map.get(aid, 0.0), period_map.get(aid, 0.0)) for aid in asset_ids}
                        
                        # Second pass for VNC filter if not cached
                        new_rows = []
                        for r in asset_rows:
                            aid = r["id_immo"]
                            real_cum = date_lines.get(aid, (0.0, 0.0))[0]
                            pv = float(r["purchase_value"] or 0)
                            is_close = r.get("state") == "close"
                            is_fully_amortized = is_close or (real_cum >= pv - 0.01)

                            if vnc_filter == "positive" or non_amortie:
                                if not is_fully_amortized:
                                    new_rows.append(r)
                            elif vnc_filter == "zero":
                                if is_fully_amortized:
                                    new_rows.append(r)
                            else:
                                new_rows.append(r)
                        asset_rows = new_rows
                        asset_ids = [r["id_immo"] for r in asset_rows]
                else:
                    date_lines = {}
                lines = []  # not needed for evolution in date mode
            else:
                # Aggregate real amount per category directly in SQL (fast, no Python loop)
                sc2 = "AND societe_comptable.name = %s" if societe else ""
                if yd_type == "date":
                    yf2 = f"AND {date_col} = %s"
                elif yd_type == "year":
                    yf2 = f"AND EXTRACT(YEAR FROM {date_col}) = %s"
                else:
                    yf2 = ""
                with conn.cursor() as cur:
                    cur.execute(f"""
                        SELECT aaat.id AS id_immo,
                               aac.id AS cat_id, 
                               COALESCE(societe_comptable.name, 'Non défini') AS company_name, 
                               SUM(aadl.amount) AS total_real
                        FROM account_asset_depreciation_line aadl
                        JOIN account_asset_asset aaat ON aaat.id = aadl.asset_id
                        JOIN account_asset_category aac ON aac.id = aaat.category_id
                        LEFT JOIN res_company societe_comptable ON societe_comptable.id = aaat.company_id
                        WHERE (aac.id = ANY(%s) {v_clause}) {sc2} {yf2}
                        GROUP BY aaat.id, aac.id, societe_comptable.name
                    """, [all_cat_ids] + ([societe] if societe else []) + ([yd_val] if yd_val is not None else []))
                    
                    real_by_cat_id = {}
                    real_by_company = {}
                    real_by_comp_cat_id = {}
                    v_map, _ = _get_virtual_category_mapping(all_cats)
                    excel_map = _get_excel_category_overrides()
                    for r in cur.fetchall():
                        aid, cid, comp, amt = r[0], r[1], r[2], float(r[3] or 0)
                        # Apply overrides: Excel first, then virtual
                        if aid in excel_map:
                            cid = excel_map[aid]
                        elif aid in v_map:
                            cid = v_map[aid]
                        
                        if cid not in all_cat_ids:
                            continue
                        real_by_cat_id[cid] = real_by_cat_id.get(cid, 0.0) + amt
                        real_by_company[comp] = real_by_company.get(comp, 0.0) + amt
                        real_by_comp_cat_id[(comp, cid)] = real_by_comp_cat_id.get((comp, cid), 0.0) + amt
                date_lines = {}
                lines = []


        # (Evolution and acquisition queries removed as requested)

        conn.close()

        # ── 5. Compute gap by category ────────────────────────────────────────

        gap_by_cat = {}       # label -> {real, estimated, count, total_pv, period_real, period_est}
        gap_by_company = {}   # company_name -> {real, estimated, count, total_pv, period_real, period_est}
        gap_by_comp_cat = {}  # (company_name, label) -> {count, total_pv, real, estimated, period_real, period_est}
        asset_details_list = []  # per-asset detail for drill-down

        if target_date:
            # Per-asset: look up real from date_lines, compute estimated via schedule
            for asset in asset_rows:
                cid   = asset["cat_id"]
                comp  = asset["company_name"]
                info  = cat_lookup.get(cid, {"label": "?", "years": 10})
                label = info["label"]
                
                # Use the NEW configured duration for the Estimated calculation (to match generator)
                years = info["years"]
                
                gap_by_cat.setdefault(label, {"real": 0.0, "estimated": 0.0, "count": 0, "total_pv": 0.0, "period_real": 0.0, "period_est": 0.0})
                gap_by_company.setdefault(comp, {"real": 0.0, "estimated": 0.0, "count": 0, "total_pv": 0.0, "period_real": 0.0, "period_est": 0.0})
                gap_by_comp_cat.setdefault((comp, label), {"count": 0, "total_pv": 0.0, "real": 0.0, "estimated": 0.0, "period_real": 0.0, "period_est": 0.0})
                
                pv = float(asset.get("purchase_value") or 0)
                rl_data = date_lines.get(asset["id_immo"], (0.0, 0.0))
                rl = rl_data[0]
                rl_amount = rl_data[1]
                
                gap_by_cat[label]["count"] += 1
                gap_by_cat[label]["total_pv"] += pv
                gap_by_cat[label]["real"] += rl
                gap_by_cat[label]["period_real"] += rl_amount
                
                gap_by_company[comp]["count"] += 1
                gap_by_company[comp]["total_pv"] += pv
                gap_by_company[comp]["real"] += rl
                gap_by_company[comp]["period_real"] += rl_amount
                
                gap_by_comp_cat[(comp, label)]["count"] += 1
                gap_by_comp_cat[(comp, label)]["total_pv"] += pv
                gap_by_comp_cat[(comp, label)]["real"] += rl
                gap_by_comp_cat[(comp, label)]["period_real"] += rl_amount
                
                # Fast estimated calculation — use years*365 fixed denominator (matches generator)
                est = 0.0
                period_est = 0.0
                start_d = asset.get("start_date") or asset.get("purchase_date")
                if start_d and pv > 0 and years > 0:
                    if isinstance(start_d, str):
                        start_d = dt_cls.strptime(start_d[:10], "%Y-%m-%d").date()
                    elif isinstance(start_d, dt_cls):
                        start_d = start_d.date()

                    from dateutil.relativedelta import relativedelta
                    from datetime import timedelta
                    # Use months to avoid ValueError with non-integer years in relativedelta
                    end_d = start_d + relativedelta(months=int(round(years*12))) - timedelta(days=1)

                    if target_date >= end_d:
                        est = pv
                    elif target_date > start_d:
                        elapsed_days = (target_date - start_d).days + 1
                        est = pv * elapsed_days / (years * 365)  # fixed denominator

                    period_start = max(start_d, first_day)
                    period_end = min(end_d, last_day)
                    if period_start <= period_end:
                        days_in_period = (period_end - period_start).days + 1
                        period_est = pv * days_in_period / (years * 365)

                gap_by_cat[label]["estimated"] += est
                gap_by_company[comp]["estimated"] += est
                gap_by_comp_cat[(comp, label)]["estimated"] += est
                gap_by_cat[label]["period_est"] += period_est
                gap_by_company[comp]["period_est"] += period_est
                gap_by_comp_cat[(comp, label)]["period_est"] += period_est

                asset_details_list.append({
                    "company":   comp,
                    "cat_id":    cid,
                    "category_label":  label,
                    "id_immo":   asset["id_immo"],
                    "code":      asset.get("code") or "",
                    "asset_name": asset.get("asset_name") or "",
                    "purchase_value":  round(pv, 2),
                    "real":      round(rl, 2),
                    "estimated": round(est, 2),
                    "gap":       round(rl - est, 2),
                    "period_real": round(rl_amount, 2),
                    "period_est":  round(period_est, 2),
                })
        else:
            # SQL-aggregated: compute estimated up to the last day of current month
            from datetime import date
            import calendar
            today = date.today()
            last_day = calendar.monthrange(today.year, today.month)[1]
            fallback_target = date(today.year, today.month, last_day)

            for cid, info in cat_lookup.items():
                label = info["label"]
                gap_by_cat.setdefault(label, {"real": 0.0, "estimated": 0.0, "count": 0, "total_pv": 0.0})
            
            for asset in asset_rows:
                cid   = asset["cat_id"]
                comp  = asset["company_name"]
                info  = cat_lookup.get(cid, {"label": "?", "years": 10})
                label = info["label"]
                years = info["years"]
                
                gap_by_company.setdefault(comp, {"real": 0.0, "estimated": 0.0, "count": 0, "total_pv": 0.0})
                gap_by_comp_cat.setdefault((comp, label), {"count": 0, "total_pv": 0.0, "real": 0.0, "estimated": 0.0})
                
                pv = float(asset.get("purchase_value") or 0)
                
                gap_by_cat[label]["count"] += 1
                gap_by_cat[label]["total_pv"] += pv
                
                gap_by_company[comp]["count"] += 1
                gap_by_company[comp]["total_pv"] += pv
                
                gap_by_comp_cat[(comp, label)]["count"] += 1
                gap_by_comp_cat[(comp, label)]["total_pv"] += pv

                # When no date is selected, estimated depreciation represents the full life schedule (100% of PV), matching generator
                est = 0.0
                start_d = asset.get("start_date") or asset.get("purchase_date")
                if start_d and pv > 0 and years > 0:
                    est = pv

                gap_by_cat[label]["estimated"] += est
                gap_by_company[comp]["estimated"] += est
                gap_by_comp_cat[(comp, label)]["estimated"] += est

                asset_details_list.append({
                    "company":   comp,
                    "cat_id":    cid,
                    "category_label":  label,
                    "id_immo":   asset["id_immo"],
                    "code":      asset.get("code") or "",
                    "asset_name": asset.get("asset_name") or "",
                    "purchase_value":  round(pv, 2),
                    "real":      None,   # not available per-asset without a target date
                    "estimated": round(est, 2),
                    "gap":       None,
                    "period_real": None,
                    "period_est":  None,
                })

            for cid, info in cat_lookup.items():
                label = info["label"]
                gap_by_cat[label]["real"] = real_by_cat_id.get(cid, 0.0)
                
            for comp, stats in gap_by_company.items():
                stats["real"] = real_by_company.get(comp, 0.0)

            for (comp, label), stats in gap_by_comp_cat.items():
                cid = None
                for cat_id, info in cat_lookup.items():
                    if info["label"] == label:
                        cid = cat_id
                        break
                if cid:
                    stats["real"] = real_by_comp_cat_id.get((comp, cid), 0.0)


        gap_list = []
        total_real = total_est = total_assets = 0

        # Build a reverse label->cid map for old/new year lookup
        label_to_cid = {}
        for cid, info in cat_lookup.items():
            label_to_cid[info["label"]] = cid

        for lbl, d in gap_by_cat.items():
            if d["count"] == 0:
                continue
            g = d["real"] - d["estimated"]
            cid      = label_to_cid.get(lbl)
            info     = cat_lookup.get(cid, {})
            new_yrs  = info.get("years")
            old_yrs  = info.get("old_years")

            total_pv   = d.get("total_pv", 0.0)
            if target_date:
                period_old = d.get("period_real")
                period_new = d.get("period_est")
            else:
                annual_old = round(total_pv / old_yrs, 2) if old_yrs and old_yrs > 0 else None
                annual_new = round(total_pv / new_yrs, 2) if new_yrs and new_yrs > 0 else None
                period_old = round(annual_old / 12, 2) if annual_old is not None else None
                period_new = round(annual_new / 12, 2) if annual_new is not None else None

            saving_per_period = round(period_old - period_new, 2) if period_old is not None and period_new is not None else None

            gap_list.append({
                "cat_id":            cid,
                "category":          lbl,
                "real":              round(d["real"], 2),
                "estimated":         round(d["estimated"], 2),
                "gap":               round(g, 2),
                "count":             d["count"],
                "old_years":         old_yrs,
                "new_years":         new_yrs,
                "total_pv":          round(total_pv, 2),
                "period_old":        round(period_old, 2) if period_old is not None else None,
                "period_new":        round(period_new, 2) if period_new is not None else None,
                "saving_per_period": saving_per_period,
            })
            total_real   += d["real"]
            total_est    += d["estimated"]
            total_assets += d["count"]
        gap_list.sort(key=lambda x: x["real"], reverse=True)
        
        # Build economy details list
        economy_list = []
        for (comp, lbl), d in gap_by_comp_cat.items():
            if d["count"] == 0:
                continue
            cid      = label_to_cid.get(lbl)
            info     = cat_lookup.get(cid, {})
            new_yrs  = info.get("years")
            old_yrs  = info.get("old_years")
            total_pv = d.get("total_pv", 0.0)
            real_val = d.get("real", 0.0)
            est_val  = d.get("estimated", 0.0)
            gap_val  = real_val - est_val
            
            if target_date:
                period_old = d.get("period_real", 0.0)
                period_new = d.get("period_est", 0.0)
            else:
                annual_old = round(total_pv / old_yrs, 2) if old_yrs and old_yrs > 0 else None
                annual_new = round(total_pv / new_yrs, 2) if new_yrs and new_yrs > 0 else None
                period_old = round(annual_old / 12, 2) if annual_old is not None else None
                period_new = round(annual_new / 12, 2) if annual_new is not None else None
            
            economy_list.append({
                "company": comp,
                "cat_id": cid,
                "category": lbl,
                "count": d["count"],
                "total_pv": round(total_pv, 2),
                "real": round(real_val, 2),
                "estimated": round(est_val, 2),
                "gap": round(gap_val, 2),
                "period_old": period_old,
                "period_new": period_new
            })
        # Sort by gap descending (or absolute gap)
        economy_list.sort(key=lambda x: x["gap"], reverse=True)
        
        # Build gap_company_list
        gap_company_list = []
        for comp, d in gap_by_company.items():
            if d["count"] == 0:
                continue
            if target_date:
                c_p_old = d.get("period_real", 0.0)
                c_p_new = d.get("period_est", 0.0)
            else:
                c_p_old = sum(x["period_old"] for x in economy_list if x["company"] == comp and x["period_old"] is not None)
                c_p_new = sum(x["period_new"] for x in economy_list if x["company"] == comp and x["period_new"] is not None)

            gap_company_list.append({
                "company": comp,
                "real": round(d["real"], 2),
                "estimated": round(d["estimated"], 2),
                "gap": round(d["real"] - d["estimated"], 2),
                "count": d["count"],
                "total_pv": round(d["total_pv"], 2),
                "period_old": round(c_p_old, 2),
                "period_new": round(c_p_new, 2)
            })
        gap_company_list.sort(key=lambda x: x["real"], reverse=True)

        # (Acq and Evo processing removed)
        all_labels = [cat_lookup[cid]["label"] for cid in sorted(cat_lookup)]
        
        # Calculate overall purchase value
        total_purchase_value = sum(d.get("total_pv", 0) for d in gap_list)

        if request.args.get("export") in ["excel", "csv"]:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            import io
            from flask import send_file

            wb = Workbook()
            ws = wb.active
            ws.title = "Analyse"
            ws.views.sheetView[0].showGridLines = True

            # Styles
            font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
            font_subtitle = Font(name="Segoe UI", size=9, italic=True, color="E0E7FF")
            font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
            font_data = Font(name="Segoe UI", size=10)
            font_total = Font(name="Segoe UI", size=10, bold=True)

            fill_title = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid") # Dark indigo
            fill_header = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo accent
            fill_total = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid") # Cool gray 100
            
            # Subtle borders
            thin_border = Border(
                left=Side(style='thin', color='E5E7EB'),
                right=Side(style='thin', color='E5E7EB'),
                top=Side(style='thin', color='E5E7EB'),
                bottom=Side(style='thin', color='E5E7EB')
            )
            total_border = Border(
                top=Side(style='thin', color='9CA3AF'),
                bottom=Side(style='double', color='111827')
            )

            align_left = Alignment(horizontal="left", vertical="center")
            align_right = Alignment(horizontal="right", vertical="center")
            align_center = Alignment(horizontal="center", vertical="center")

            # Determine if this is a detailed or summary export
            societe_val = request.args.get("societe")
            if societe_val in ["Toutes", "undefined", "null", "all", ""]:
                societe_val = None

            if societe_val:
                # DETAILED ASSETS EXPORT
                filename = f"analyse_details_{societe_val.replace(' ', '_')}.xlsx"
                title_text = f"Analyse Détaillée des Écarts d'Amortissement — {societe_val}"
                
                # Title Block
                ws.merge_cells("A1:L2")
                title_cell = ws["A1"]
                title_cell.value = title_text
                title_cell.font = font_title
                title_cell.fill = fill_title
                title_cell.alignment = align_center

                # Subtitle Block (parameters)
                ws.merge_cells("A3:L3")
                sub_cell = ws["A3"]
                date_param = request.args.get("date") or "Toutes dates"
                gran_param = request.args.get("gran") or "mensuelle"
                sub_cell.value = f"Date cible: {date_param} | Granularité: {gran_param} | Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                sub_cell.font = font_subtitle
                sub_cell.fill = fill_title
                sub_cell.alignment = align_center

                headers = [
                    "Catégorie", "ID Catégorie", "ID Immo", "Code", "Désignation", 
                    "Valeur Brute", "Amorti Réel", "Amorti Estimé", 
                    "Écart (Réel - Estimé)", "Écart Période (Réel - Estimé)",
                    "Dotation Période Réelle", "Dotation Période Estimée"
                ]
                
                # Set row heights
                ws.row_dimensions[1].height = 18
                ws.row_dimensions[2].height = 18
                ws.row_dimensions[3].height = 16
                ws.row_dimensions[5].height = 24 # Header row

                # Write headers
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=5, column=col_idx, value=header)
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = align_center
                    cell.border = thin_border

                # Write data
                current_row = 6
                total_pv = 0.0
                total_real = 0.0
                total_est = 0.0
                total_gap = 0.0
                total_p_real = 0.0
                total_p_est = 0.0

                for asset in asset_details_list:
                    # Filter just in case
                    if asset.get("company") != societe_val:
                        continue
                        
                    pv = asset.get("purchase_value") or 0.0
                    real = asset.get("real") or 0.0
                    est = asset.get("estimated") or 0.0
                    gap = (real - est) if (real is not None and est is not None) else (asset.get("gap") or 0.0)
                    p_real = asset.get("period_real") or 0.0
                    p_est = asset.get("period_est") or 0.0

                    total_pv += pv
                    total_real += real
                    total_est += est
                    total_gap += gap
                    total_p_real += p_real
                    total_p_est += p_est

                    row_vals = [
                        asset.get("category_label") or "Sans catégorie",
                        asset.get("cat_id"),
                        asset.get("id_immo"),
                        asset.get("code") or "",
                        asset.get("asset_name") or "",
                        pv,
                        real,
                        est,
                        gap,
                        p_real - p_est,
                        p_real,
                        p_est
                    ]

                    for col_idx, val in enumerate(row_vals, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=val)
                        cell.font = font_data
                        cell.border = thin_border
                        
                        # Formatting and alignment
                        if col_idx in [1, 4, 5]:
                            cell.alignment = align_left
                        elif col_idx in [2, 3]:
                            cell.alignment = align_center
                        else:
                            cell.alignment = align_right
                            cell.number_format = '#,##0.00'

                    current_row += 1

                # Write Total Row
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                total_label = ws.cell(row=current_row, column=1, value="TOTAL SÉLECTION")
                total_label.font = font_total
                total_label.alignment = align_left
                total_label.fill = fill_total

                # Fill standard cell styles for merged region borders
                for c in range(1, 6):
                    ws.cell(row=current_row, column=c).border = total_border
                    ws.cell(row=current_row, column=c).fill = fill_total

                totals = {
                    6: total_pv,
                    7: total_real,
                    8: total_est,
                    9: total_gap,
                    10: total_p_real - total_p_est,
                    11: total_p_real,
                    12: total_p_est
                }

                for col_idx, val in totals.items():
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.font = font_total
                    cell.fill = fill_total
                    cell.alignment = align_right
                    cell.number_format = '#,##0.00'
                    cell.border = total_border

            else:
                # SUMMARY ECONOMY EXPORT
                filename = "analyse_ecart_amortissement.xlsx"
                title_text = "Synthèse Globale des Écarts d'Amortissement"

                # Title Block
                ws.merge_cells("A1:K2")
                title_cell = ws["A1"]
                title_cell.value = title_text
                title_cell.font = font_title
                title_cell.fill = fill_title
                title_cell.alignment = align_center

                # Subtitle Block (parameters)
                ws.merge_cells("A3:K3")
                sub_cell = ws["A3"]
                date_param = request.args.get("date") or "Toutes dates"
                gran_param = request.args.get("gran") or "mensuelle"
                sub_cell.value = f"Date cible: {date_param} | Granularité: {gran_param} | Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                sub_cell.font = font_subtitle
                sub_cell.fill = fill_title
                sub_cell.alignment = align_center

                headers = [
                    "Société", "Catégorie", "ID Catégorie", "Nombre d'Immo", "Valeur Brute", 
                    "Amorti Réel", "Amorti Estimé", "Écart (Charge/Produit)", 
                    "Écart Période (Charge/Produit)",
                    "Dotation Période Réelle", "Dotation Période Estimée"
                ]

                # Set row heights
                ws.row_dimensions[1].height = 18
                ws.row_dimensions[2].height = 18
                ws.row_dimensions[3].height = 16
                ws.row_dimensions[5].height = 24 # Header row

                # Write headers
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=5, column=col_idx, value=header)
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = align_center
                    cell.border = thin_border

                # Write data
                current_row = 6
                total_count = 0
                total_pv = 0.0
                total_real = 0.0
                total_est = 0.0
                total_gap = 0.0
                total_p_real = 0.0
                total_p_est = 0.0

                for row in economy_list:
                    count = row.get('count') or 0
                    pv = row.get('total_pv') or 0.0
                    real = row.get('real') or 0.0
                    est = row.get('estimated') or 0.0
                    gap = row.get('gap') or (real - est)
                    p_real = row.get('period_old') or 0.0
                    p_est = row.get('period_new') or 0.0

                    total_count += count
                    total_pv += pv
                    total_real += real
                    total_est += est
                    total_gap += gap
                    total_p_real += p_real
                    total_p_est += p_est

                    row_vals = [
                        row.get('company') or "",
                        row.get('category') or "",
                        row.get('cat_id'),
                        count,
                        pv,
                        real,
                        est,
                        gap,
                        p_real - p_est,
                        p_real,
                        p_est
                    ]

                    for col_idx, val in enumerate(row_vals, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=val)
                        cell.font = font_data
                        cell.border = thin_border
                        
                        # Formatting and alignment
                        if col_idx in [1, 2]:
                            cell.alignment = align_left
                        elif col_idx in [3, 4]:
                            cell.alignment = align_center
                            if col_idx == 4: cell.number_format = '#,##0'
                        else:
                            cell.alignment = align_right
                            cell.number_format = '#,##0.00'

                    current_row += 1

                # Write Total Row
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
                total_label = ws.cell(row=current_row, column=1, value="TOTAL SYNTHÈSE")
                total_label.font = font_total
                total_label.alignment = align_left
                total_label.fill = fill_total

                # Fill standard cell styles for merged region borders
                for c in range(1, 4):
                    ws.cell(row=current_row, column=c).border = total_border
                    ws.cell(row=current_row, column=c).fill = fill_total

                totals = {
                    4: total_count,
                    5: total_pv,
                    6: total_real,
                    7: total_est,
                    8: total_gap,
                    9: total_p_real - total_p_est,
                    10: total_p_real,
                    11: total_p_est
                }

                for col_idx, val in totals.items():
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.font = font_total
                    cell.fill = fill_total
                    cell.alignment = align_right if col_idx != 4 else align_center
                    cell.number_format = '#,##0.00' if col_idx != 4 else '#,##0'
                    cell.border = total_border

            # Auto-adjust column widths dynamically
            for col in ws.columns:
                max_len = 0
                for cell in col:
                    # Skip merged title and subtitle cell length calculations
                    if cell.row in [1, 2, 3]:
                        continue
                    if cell.value is not None:
                        val_str = str(cell.value)
                        # Add padding for currency formatting or numbers
                        if isinstance(cell.value, (int, float)):
                            val_str = f"{cell.value:,.2f}"
                        max_len = max(max_len, len(val_str))
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            # Save in-memory and send
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)

            return send_file(
                out,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=filename
            )

        return jsonify({
            "gap_by_category": gap_list,
            "gap_by_company": gap_company_list,
            "economy_details": economy_list,
            "asset_details": asset_details_list,
            "all_labels": all_labels,
            "date_requested": date_str is not None,
            "period_info": {
                "start": first_day_str if date_str else None,
                "end": last_day_str if date_str else None,
                "granularity": eff_gran if date_str else None
            },
            "kpis": {
                "total_real":      round(total_real, 2),
                "total_estimated": round(total_est, 2),
                "total_gap":       round(total_real - total_est, 2),
                "total_assets":    total_assets,
                "total_pv":        round(total_purchase_value, 2)
            }
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/api/generate", methods=["POST"])
def api_generate():
    """Start generation in background thread.

    Supports two calling conventions:
      1. Legacy (polling):  {categories, output_format, asset_id, societe, dep_date}
         → returns {ok: true}, poll /api/status
      2. SSE (index page):  {categories, format, granularity, asset_id, from_date}
         → returns {job_id}, stream /api/stream/<job_id>
    """
    with _lock:
        if _job["status"] == "running":
            return jsonify({"error": "Une generation est deja en cours."}), 409

    body = request.get_json(force=True) or {}

    # ── Detect calling convention ─────────────────────────────────────────────
    is_sse = "format" in body or "granularity" in body or "from_date" in body

    # Common fields
    categories   = body.get("categories", [])
    asset_id_raw = body.get("asset_id") or None
    asset_id     = int(asset_id_raw) if asset_id_raw and str(asset_id_raw).isdigit() else None
    asset_search = (body.get("asset_search") or "").strip() or None
    year_debut = body.get("year_debut") or None
    year_debut_type = body.get("year_debut_type") or "start_date"
    societe      = body.get("societe") or None
    ignore_overrides = body.get("ignore_overrides", False)
    vnc_filter   = body.get("vnc_filter", "all")

    if not categories:
        return jsonify({"error": "Aucune categorie selectionnee."}), 400

    if is_sse:
        # ── SSE mode (index.html) ─────────────────────────────────────────────
        output_format = body.get("format", "excel")
        granularity   = body.get("granularity", "monthly")
        from_date_raw = (body.get("from_date") or "").strip()
        as_of_raw     = (body.get("as_of_date") or "").strip()

        _, from_date = _parse_flexible_date(from_date_raw)
        if from_date_raw and not from_date:
            return jsonify({"error": f"Date invalide '{from_date_raw}' — format AAAA, MM/AAAA ou JJ/MM/AAAA requis."}), 400

        dep_date_str, as_of_date = _parse_flexible_date(as_of_raw)
        if as_of_raw and not as_of_date:
            return jsonify({"error": f"Date invalide '{as_of_raw}' — format AAAA, MM/AAAA ou JJ/MM/AAAA requis."}), 400

        # Resolve category labels → full dicts from config
        cfg = _load_config()
        all_cats = cfg.get("categories", [])
        label_set = set(categories)
        resolved  = [c for c in all_cats if c.get("label") in label_set]
        if not resolved:
            return jsonify({"error": "Aucune categorie valide."}), 400

        job_id = str(uuid.uuid4())
        log_q: queue.Queue = queue.Queue()
        with _sse_lock:
            _sse_jobs[job_id] = {"status": "running", "log_q": log_q, "file": None}

        thread = threading.Thread(
            target=_run_generation,
            kwargs=dict(
                categories=resolved,
                output_format=output_format,
                granularity=granularity,
                from_date=from_date,
                as_of_date=as_of_date,
                dep_date=dep_date_str,
                asset_id=asset_id,
                asset_search=asset_search,
                sse_job_id=job_id,
                year_debut=year_debut,
                year_debut_type=year_debut_type,
                societe=societe,
                non_amortie=body.get("non_amortie", False),
                ignore_overrides=ignore_overrides,
                vnc_filter=vnc_filter,
            ),
            daemon=True,
        )
        thread.start()
        return jsonify({"job_id": job_id})

    else:
        # ── Legacy polling mode (dashboard) ───────────────────────────────────
        output_format = body.get("output_format", "excel")
        dep_date_raw = body.get("dep_date") or None
        dep_date_str, as_of_date = _parse_flexible_date(dep_date_raw)
        if dep_date_raw and not as_of_date:
            return jsonify({"error": f"Date invalide '{dep_date_raw}' — format AAAA, MM/AAAA ou JJ/MM/AAAA requis."}), 400

        # Resolve category labels → full dicts from config (legacy mode)
        cfg = _load_config()
        all_cats = cfg.get("categories", [])
        label_set = set(categories)
        resolved  = [c for c in all_cats if c.get("label") in label_set]
        if not resolved:
            # Fallback: if they are already IDs or if resolution fails, keep as is
            # but usually they should be labels from the UI.
            resolved = categories

        thread = threading.Thread(
            target=_run_generation,
            kwargs=dict(
                categories=resolved,
                output_format=output_format,
                asset_id=asset_id,
                societe=societe,
                dep_date=dep_date_str,
                as_of_date=as_of_date,
                year_debut=year_debut,
                year_debut_type=year_debut_type,
                non_amortie=body.get("non_amortie", False),
                ignore_overrides=ignore_overrides,
                vnc_filter=vnc_filter,
            ),
            daemon=True,
        )
        thread.start()
        return jsonify({"ok": True})


@app.route("/api/status")
def api_status():
    """Return current job status + logs (legacy polling used by dashboard)."""
    with _lock:
        return jsonify(dict(_job))


@app.route("/api/stream/<job_id>")
def api_stream(job_id):
    """Server-Sent Events: stream log lines for a running SSE job."""
    with _sse_lock:
        job = _sse_jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job introuvable."}), 404

    log_q = job["log_q"]

    def generate_events():
        while True:
            try:
                msg = log_q.get(timeout=30)
            except queue.Empty:
                yield "event: keepalive\ndata: \n\n"
                continue

            if msg is None:           # sentinel — job finished
                with _sse_lock:
                    j = _sse_jobs.get(job_id, {})
                status = j.get("status", "error")
                fp     = j.get("file") or ""
                yield f"event: done\ndata: {status}|{fp}\n\n"
                break

            safe = msg.replace("\n", " ")
            yield f"data: {safe}\n\n"

    return Response(
        stream_with_context(generate_events()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.route("/api/download")
def api_download():
    """Download via legacy global job (dashboard)."""
    with _lock:
        filepath = _job.get("output_file")
    if not filepath or not os.path.exists(filepath):
        return jsonify({"error": "Fichier non disponible."}), 404
    return send_file(filepath, as_attachment=True)


@app.route("/api/download/<job_id>")
def api_download_sse(job_id):
    """Download via SSE job ID (index.html generator page)."""
    with _sse_lock:
        job = _sse_jobs.get(job_id)
    if not job or not job.get("file"):
        return jsonify({"error": "Fichier non disponible."}), 404
    fp = job["file"]
    if not os.path.exists(fp):
        return jsonify({"error": "Fichier introuvable sur le disque."}), 404
    return send_file(fp, as_attachment=True)


@app.route("/api/refresh_cache", methods=["POST"])
def api_refresh_cache():
    if global_cache.is_loading:
        return jsonify({"status": "loading", "message": "Actualisation déjà en cours..."})
    threading.Thread(target=global_cache.refresh, args=(_load_config(),), daemon=True).start()
    return jsonify({"status": "started", "message": "Actualisation lancée en arrière-plan."})


@app.route("/api/cache_status")
def api_cache_status():
    with global_cache.lock:
        return jsonify({
            "is_loading": global_cache.is_loading,
            "last_update": global_cache.last_update.strftime("%d/%m/%Y %H:%M:%S") if global_cache.last_update else "Jamais",
            "assets_count": len(global_cache.assets),
            "error": global_cache.error
        })



@app.route("/medina_analyse")
def medina_analyse():
    return render_template("medina_analyse.html")


@app.route("/api/medina_upload", methods=["POST"])
def api_medina_upload():
    if "file" not in request.files:
        return jsonify({"error": "Aucun fichier fourni."}), 400
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "Nom de fichier vide."}), 400
    
    save_path = os.path.join(SCRIPT_DIR, "medina_analyse.xlsx")
    try:
        file.save(save_path)
        # Clear the generator date override cache
        import generator
        generator._excel_date_override_cache = None
        return jsonify({"ok": True, "message": "Fichier medina_analyse.xlsx téléversé avec succès."})
    except Exception as e:
        return jsonify({"error": f"Erreur lors de la sauvegarde : {str(e)}"}), 500


@app.route("/api/medina_generate", methods=["POST"])
def api_medina_generate():
    """Start Medina generation in background thread."""
    with _lock:
        if _job["status"] == "running":
            return jsonify({"error": "Une generation est deja en cours."}), 409

    body = request.get_json(force=True) or {}
    categories = body.get("categories", [])
    output_format = body.get("format", "excel")
    granularity = body.get("granularity", "monthly")
    ignore_overrides = body.get("ignore_overrides", False)
    vnc_filter = body.get("vnc_filter", "all")
    
    asset_id_raw = body.get("asset_id") or None
    asset_id = int(asset_id_raw) if asset_id_raw and str(asset_id_raw).isdigit() else None
    asset_search = (body.get("asset_search") or "").strip() or None
    
    from_date_raw = (body.get("from_date") or "").strip()
    as_of_raw = (body.get("as_of_date") or "").strip()

    _, from_date = _parse_flexible_date(from_date_raw)
    if from_date_raw and not from_date:
        return jsonify({"error": f"Date invalide '{from_date_raw}' — format AAAA, MM/AAAA ou JJ/MM/AAAA requis."}), 400

    dep_date_str, as_of_date = _parse_flexible_date(as_of_raw)
    if as_of_raw and not as_of_date:
        return jsonify({"error": f"Date invalide '{as_of_raw}' — format AAAA, MM/AAAA ou JJ/MM/AAAA requis."}), 400

    if not categories:
        return jsonify({"error": "Aucune categorie selectionnee."}), 400

    # Resolve category labels → full dicts from config
    cfg = _load_config()
    all_cats = cfg.get("categories", [])
    label_set = set(categories)
    resolved = [c for c in all_cats if c.get("label") in label_set]
    if not resolved:
        return jsonify({"error": "Aucune categorie valide."}), 400

    job_id = str(uuid.uuid4())
    log_q: queue.Queue = queue.Queue()
    with _sse_lock:
        _sse_jobs[job_id] = {"status": "running", "log_q": log_q, "file": None}

    thread = threading.Thread(
        target=_run_generation,
        kwargs=dict(
            categories=resolved,
            output_format=output_format,
            granularity=granularity,
            from_date=from_date,
            as_of_date=as_of_date,
            dep_date=dep_date_str,
            asset_id=asset_id,
            asset_search=asset_search,
            sse_job_id=job_id,
            societe=None,
            non_amortie=body.get("non_amortie", False),
            is_medina=True,
            ignore_overrides=ignore_overrides,
            vnc_filter=vnc_filter,
        ),
        daemon=True,
    )
    thread.start()
    return jsonify({"job_id": job_id})


if __name__ == "__main__":
    print("=" * 55)
    print("  Asset Estimator - HASNAOUI GROUPE")
    print("  Ouvrir : http://localhost:5000")
    print("=" * 55)
    app.run(debug=True, port=5000)
