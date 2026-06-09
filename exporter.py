"""
exporter.py — Writes results to Excel with professional formatting.

One sheet per category + one combined "Tout" sheet.
"""

import os
from datetime import date
import pandas as pd
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


# ── Color palette ────────────────────────────────────────────────────────────
HEADER_BG   = "1F3864"   # dark navy
HEADER_FG   = "FFFFFF"   # white
ALT_ROW_BG  = "EBF0FA"   # light blue-gray
BORDER_CLR  = "C0C8D8"

import re

_ILLEGAL_XML_CHARS_RE = re.compile(
    r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x84\x86-\x9f\ud800-\udfff\ufffe\uffff]"
)

def _clean_illegal_xml_chars(val):
    if not isinstance(val, str):
        return val
    return _ILLEGAL_XML_CHARS_RE.sub("", val)


def _style_sheet(ws, df: pd.DataFrame):
    """Apply header style, alternating rows, column widths to a worksheet."""
    import numpy as np

    # Enable gridlines
    ws.views.sheetView[0].showGridLines = True

    # ── Borders ──────────────────────────────────────────────────────────────
    thin_side = Side(style="thin", color="E5E7EB")
    border_data = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    header_side = Side(style="thin", color="E5E7EB")
    border_header = Border(left=header_side, right=header_side, top=header_side, bottom=header_side)

    # ── Styles ───────────────────────────────────────────────────────────────
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Premium Indigo
    align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

    font_data = Font(name="Segoe UI", size=10)
    fill_alt = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") # Off-white alternating rows
    
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")

    # Header styling
    ws.row_dimensions[1].height = 24
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header

    ws.freeze_panes = "A2"

    # Pre-compute column properties once for performance
    col_styles = []
    for col_name in df.columns:
        col_series = df[col_name]
        is_numeric = pd.api.types.is_numeric_dtype(col_series)
        is_id_or_code = "id" in col_name.lower() or "code" in col_name.lower() or "compte" in col_name.lower() or "ref" in col_name.lower()
        
        if is_numeric and not is_id_or_code:
            align = align_right
            num_fmt = '#,##0' if pd.api.types.is_integer_dtype(col_series) else '#,##0.00'
        elif is_id_or_code or col_name.lower() in ["date", "période", "statut"]:
            align = align_center
            num_fmt = None
        else:
            align = align_left
            num_fmt = None
            
        col_styles.append((align, num_fmt))

    # Data styling
    num_rows = len(df)
    if num_rows > 5000:
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            col_series = df[col_name].head(1000)
            if len(df) and pd.api.types.is_numeric_dtype(col_series):
                max_data_len = col_series.map(lambda x: len(f"{x:,.2f}") if isinstance(x, (int, float)) else len(str(x))).max()
            else:
                max_data_len = col_series.astype(str).str.len().max() if len(df) else 0
            width = min(max(len(str(col_name)), max_data_len, 10) + 4, 45)
            ws.column_dimensions[col_letter].width = width
        return

    for row_idx in range(2, num_rows + 2):
        ws.row_dimensions[row_idx].height = 18
        
        # Alternating row fill
        use_alt = (row_idx % 2 == 1)
        
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font_data
            cell.border = border_data
            if use_alt:
                cell.fill = fill_alt
                
            align, num_fmt = col_styles[col_idx - 1]
            cell.alignment = align
            if num_fmt:
                cell.number_format = num_fmt

    # ── Auto column widths ───────────────────────────────────────────────────
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        # Width = max of header length vs max data length, capped at 45
        col_series = df[col_name].head(1000)
        if len(df) and pd.api.types.is_numeric_dtype(col_series):
            max_data_len = col_series.map(lambda x: len(f"{x:,.2f}") if isinstance(x, (int, float)) else len(str(x))).max()
        else:
            max_data_len = col_series.astype(str).str.len().max() if len(df) else 0
        width = min(max(len(str(col_name)), max_data_len, 10) + 4, 45)
        ws.column_dimensions[col_letter].width = width


def export_to_excel(results: dict[str, pd.DataFrame], output_dir: str):
    """
    Export category DataFrames to a single Excel file.

    Args:
        results:    {category_label: DataFrame}
        output_dir: Directory where the file will be saved.
    """
    os.makedirs(output_dir, exist_ok=True)
    today = date.today().isoformat()
    filepath = os.path.join(output_dir, f"amortissement_{today}.xlsx")

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:

        # ── One sheet per category ────────────────────────────────────────────
        for label, df in results.items():
            # Excel sheet names: max 31 chars, no [] : * ? / \
            import re
            sheet_name = re.sub(r'[\[\]\:\*\?\/\\]', '-', label[:31])
            if not sheet_name: sheet_name = "Sheet"
            
            # Clean data: only on columns that might contain illegal XML characters for speed
            df_clean = df.copy()
            for col in df_clean.columns:
                if col in ["Désignation", "Code", "Nom", "Name", "designation", "code"]:
                    if df_clean[col].dtype == object:
                        df_clean[col] = df_clean[col].map(lambda x: _clean_illegal_xml_chars(x) if isinstance(x, str) else x)
            
            df_clean.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            _style_sheet(ws, df_clean)

        # ── Combined "Tout" sheet ─────────────────────────────────────────────
        if len(results) > 1:
            combined = pd.concat(
                [df.assign(**{"Catégorie (filtre)": label})
                 for label, df in results.items()],
                ignore_index=True
            )
            combined_clean = combined.copy()
            for col in combined_clean.columns:
                if col in ["Désignation", "Code", "Nom", "Name", "designation", "code"]:
                    if combined_clean[col].dtype == object:
                        combined_clean[col] = combined_clean[col].map(lambda x: _clean_illegal_xml_chars(x) if isinstance(x, str) else x)
                    
            combined_clean.to_excel(writer, sheet_name="Tout", index=False)
            ws = writer.sheets["Tout"]
            _style_sheet(ws, combined_clean)

    print(f"\n[OK] Export termine : {filepath}")
    return filepath


def export_to_csv(results: dict[str, pd.DataFrame], output_dir: str):
    """Export each category as a separate CSV file."""
    os.makedirs(output_dir, exist_ok=True)
    today = date.today().isoformat()
    paths = []
    for label, df in results.items():
        safe_label = label.replace("%", "").replace(" ", "_").strip("_")
        filepath = os.path.join(output_dir, f"amortissement_{safe_label}_{today}.csv")
        df.to_csv(filepath, index=False, encoding="utf-8-sig")  # utf-8-sig for Excel compat
        paths.append(filepath)
        print(f"[OK] CSV : {filepath}")
    return paths



